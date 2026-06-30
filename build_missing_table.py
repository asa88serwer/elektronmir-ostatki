"""Генерация Excel-таблицы 1000 позиций без характеристик.

Приоритет: основные security-бренды первыми, затем остальные из склада.
Для каждой позиции формируется предполагаемый URL поиска на сайте производителя.
"""

import sys, os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, ".")
from merge_warehouses import extract_all_items
from parse_specs import build_specs_map

DIR = os.path.dirname(os.path.abspath(__file__))

# Бренды в порядке приоритета
PRIORITY_BRANDS = [
    ("dahua",     "Dahua",     "https://www.dahua.com/en/products/", "dahua.com"),
    ("hikvision", "Hikvision", "https://www.hikvision.com/en/products/", "hikvision.com"),
    ("zkteco",    "ZKTeco",    "https://www.zkteco.eu/index.php?c=product&a=search&searchValue=", "zkteco.eu"),
    ("tiandy",    "Tiandy",    "https://www.tiandy.com/products/", "tiandy.com"),
    ("temid",     "Temid",     "https://temid.ru/catalog/", "temid.ru"),
    ("hiwatch",   "HiWatch",   "https://www.hiwatch.ru/catalog/", "hiwatch.ru"),
    ("ezviz",     "EZVIZ",     "https://www.ezviz.com/product/", "ezviz.com"),
    ("optimus",   "Optimus",   "https://optimus-cctv.ru/catalog/", "optimus-cctv.ru"),
    ("belpark",   "Belpark",   "https://belpark.ru/catalog/", "belpark.ru"),
]

LIMIT = 1000

def guess_category(name):
    nl = name.lower()
    checks = [
        (["nvr","dvr","xvr","регистратор","recorder"], "Регистратор"),
        (["ipc","hac-","sd2","sd3","sd4","sd5","sd6","sd8","tpc","itc","ptz","камер","camera","bullet","dome","turret","fisheye"], "Камера"),
        (["vto","vth","vtm","домофон","intercom","video door"], "Домофония"),
        (["switch","коммутатор","poe"], "Коммутатор"),
        (["ups","блок питания","power supply","psu"], "Блок питания"),
        (["dss","license","лицензи","software","vms"], "ПО/Лицензия"),
        (["inbio","c3-","asc","skd","скуд","controller","контроллер"], "Контроллер СКУД"),
        (["speedface","proface","fr","биометр","terminal","терминал","face"], "Биометр. терминал"),
        (["считыватель","reader","kr1","kr2","proid"], "Считыватель"),
        (["турникет","turnstile","ts1","ts2","fbl"], "Турникет"),
        (["шлагбаум","boom","barrier","bgm","bg1"], "Шлагбаум"),
        (["замок","lock","al-","lm-","электро"], "Замок"),
        (["bracket","pfa","pfb","кронштейн","mount","holder"], "Крепление"),
        (["card","tag","карта","метка"], "Карта/Метка"),
        (["кабель","cable","sfp","patch"], "Кабель"),
        (["монитор","monitor","display","экран"], "Монитор"),
        (["сигнализац","alarm","arc","ara","detector","датчик"], "Сигнализация"),
        (["kit","комплект"], "Комплект"),
    ]
    for keywords, cat in checks:
        if any(k in nl for k in keywords):
            return cat
    return "Прочее"


def build_search_url(brand_key, brand_site, name):
    import urllib.parse
    model = name.strip()
    if brand_key == "dahua":
        return f"https://www.dahua.com/en/search/?q={urllib.parse.quote(model)}"
    if brand_key == "hikvision":
        return f"https://www.hikvision.com/en/search/?keywords={urllib.parse.quote(model)}"
    if brand_key == "zkteco":
        return f"https://www.zkteco.eu/index.php?c=product&a=search&searchValue={urllib.parse.quote(model)}"
    return f"https://www.google.com/search?q={urllib.parse.quote(model + ' характеристики site:' + brand_site)}"


def main():
    print("Загрузка склада...")
    minsk = extract_all_items(os.path.join(DIR, "Минск.xls"))
    moscow = extract_all_items(os.path.join(DIR, "Москва.xls"))
    all_names = sorted(set(list(minsk.keys()) + list(moscow.keys())))
    print(f"Всего позиций: {len(all_names)}")

    print("Загрузка характеристик...")
    specs = build_specs_map(all_names)
    print(f"Уже есть характеристики: {len(specs)}")

    # Собираем позиции без характеристик по приоритетным брендам
    missing = []
    seen = set()

    for brand_key, brand_name, brand_url, brand_site in PRIORITY_BRANDS:
        items = [n for n in all_names
                 if n.lower().startswith(brand_key) and n not in specs and n not in seen]
        for name in items:
            seen.add(name)
            qty_minsk = int(minsk.get(name, 0))
            qty_moscow = int(moscow.get(name, 0))
            missing.append({
                "brand": brand_name,
                "name": name,
                "category": guess_category(name),
                "qty_minsk": qty_minsk,
                "qty_moscow": qty_moscow,
                "qty_total": qty_minsk + qty_moscow,
                "search_url": build_search_url(brand_key, brand_site, name),
            })

    print(f"Основные бренды (без характеристик): {len(missing)}")

    # Если меньше 1000 — добираем остальные позиции из склада (по убыванию суммарного остатка)
    if len(missing) < LIMIT:
        rest = []
        for name in all_names:
            if name in seen or name in specs:
                continue
            qty_minsk = int(minsk.get(name, 0))
            qty_moscow = int(moscow.get(name, 0))
            rest.append({
                "brand": name.split()[0],
                "name": name,
                "category": guess_category(name),
                "qty_minsk": qty_minsk,
                "qty_moscow": qty_moscow,
                "qty_total": qty_minsk + qty_moscow,
                "search_url": f"https://www.google.com/search?q={name.replace(' ', '+')}+характеристики+datasheet",
            })
        # Сначала позиции с реальными остатками
        rest.sort(key=lambda x: -x["qty_total"])
        for item in rest:
            if len(missing) >= LIMIT:
                break
            seen.add(item["name"])
            missing.append(item)

    # Сортируем: сначала с остатками > 0, затем нулевые
    missing.sort(key=lambda x: (-x["qty_total"], x["brand"], x["name"]))
    missing = missing[:LIMIT]
    print(f"Итого в таблице: {len(missing)}")

    # Сохраняем JSON для последующего парсинга
    json_path = os.path.join(DIR, "missing_1000.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(missing, f, ensure_ascii=False, indent=2)
    print(f"JSON сохранён: {json_path}")

    # Генерируем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Без характеристик"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    alt_fill = PatternFill(start_color="F5F9FF", end_color="F5F9FF", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["№", "Бренд", "Наименование", "Категория", "Минск", "Москва", "Итого", "Ссылка для поиска"]
    col_widths = [5, 12, 65, 22, 7, 7, 7, 60]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H1"

    for i, item in enumerate(missing, 1):
        row = i + 1
        fill = alt_fill if i % 2 == 0 else None
        vals = [i, item["brand"], item["name"], item["category"],
                item["qty_minsk"], item["qty_moscow"], item["qty_total"], item["search_url"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin
            if col in (5, 6, 7):
                c.alignment = Alignment(horizontal="center")
            if col == 1:
                c.alignment = Alignment(horizontal="center")
            if fill:
                c.fill = fill

    # Лист сводки по брендам
    ws2 = wb.create_sheet("Сводка")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18

    for col, h in enumerate(["Бренд", "Позиций", "С остатками", "Без остатков", "Категорий"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    from collections import Counter
    brands_in_table = Counter(x["brand"] for x in missing)
    for r, (brand, cnt) in enumerate(sorted(brands_in_table.items()), 2):
        brand_items = [x for x in missing if x["brand"] == brand]
        with_stock = sum(1 for x in brand_items if x["qty_total"] > 0)
        categories = len(set(x["category"] for x in brand_items))
        ws2.cell(row=r, column=1, value=brand).border = thin
        ws2.cell(row=r, column=2, value=cnt).border = thin
        ws2.cell(row=r, column=3, value=with_stock).border = thin
        ws2.cell(row=r, column=4, value=cnt - with_stock).border = thin
        ws2.cell(row=r, column=5, value=categories).border = thin
        for col in range(2, 6):
            ws2.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    out = os.path.join(DIR, "Без_характеристик_1000.xlsx")
    wb.save(out)
    print(f"Excel сохранён: {out}")
    return missing


if __name__ == "__main__":
    main()
