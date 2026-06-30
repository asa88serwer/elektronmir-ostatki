import sys, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, ".")
from parse_specs import build_specs_map
from merge_warehouses import extract_all_items

all_names = sorted(set(list(extract_all_items("Минск.xls").keys()) + list(extract_all_items("Москва.xls").keys())))
specs = build_specs_map(all_names)

brands_tabs = ["dahua","hikvision","hiwatch","tiandy","zkteco","temid","ezviz","optimus","belpark"]
missing = {}
for name in all_names:
    nl = name.lower()
    for b in brands_tabs:
        if nl.startswith(b):
            if name not in specs:
                missing.setdefault(b, []).append(name)
            break

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Без характеристик"

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
brand_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
brand_font = Font(bold=True, size=11, color="0D47A1")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

headers = ["№", "Бренд", "Артикул / Наименование", "Категория (предполаг.)"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 65
ws.column_dimensions["D"].width = 30
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:D1"

brand_display = {
    "dahua": "Dahua", "hikvision": "Hikvision", "hiwatch": "HiWatch",
    "tiandy": "Tiandy", "zkteco": "ZKTeco", "temid": "Temid",
    "ezviz": "EZVIZ", "optimus": "Optimus", "belpark": "Belpark",
}


def guess_category(nl):
    checks = [
        (["lm2", "lm4", "lm7", "ls5", "монитор", "display"], "Монитор"),
        (["vto", "vth", "vtm", "vtob", "vtof", "ktp", "ktw", "kis", "vdp"], "Домофония"),
        (["pfa", "pfb", "pfm", "pft", "bracket"], "Аксессуар/Крепление"),
        (["ara", "ard", "arc", "arm", "ark", "ics"], "Охранная сигнализация"),
        (["dss", "license", "лицензи"], "ПО/Лицензия"),
        (["as5", "sg5", "s4110", "switch"], "Коммутатор"),
        (["xvr", "nvr", "dvr"], "Регистратор"),
        (["ipc", "hac-h", "hac-b", "hac-d", "hac-t", "sd2c", "sd3", "sd4", "sd5", "sd6", "sd8", "sdt", "tpc", "itc", "ptz-"], "Камера"),
        (["al-1", "al-2", "al-3", "al-4", "lm-", "lmb-", "cm-2"], "Замок/Электрозащёлка"),
        (["inbio", "c3-", "asc"], "Контроллер СКУД"),
        (["kr1", "kr2", "proid", "mr1", "mk-v", "skw", "cr2", "fr1"], "Считыватель"),
        (["ts1", "ts2", "ts5", "ts8", "tsa", "fbl", "tfg"], "Турникет/Проходная"),
        (["bgm", "bg1", "sbt", "cmp", "boom"], "Шлагбаум/Стрела"),
        (["speedface", "proface", "iclock", "sf3", "sf4", "sensefac"], "Биометрический терминал"),
        (["card", "tag"], "Карта/Метка"),
        (["ex-8", "ex-8", "кнопк"], "Кнопка выхода"),
        (["eb1", "eb3", "eb10", "dc40", "dc80", "edf", "drl"], "Блок питания"),
        (["sfp", "gsfp", "1ln", "hk-"], "Кабель/SFP"),
        (["kit", "комплект"], "Комплект"),
        (["act-", "hwc-"], "Камера/Аксессуар"),
        (["isi", "asi", "asr", "asa", "ds0"], "СКУД/Терминал"),
        (["isc-", "ipme"], "Оборудование"),
        (["esl-", "hy-", "nkb", "pkp", "mlc", "ivd"], "Оборудование"),
        (["sa4", "sa3", "x8s", "zk-d", "wl", "wr4", "ll-", "grm", "d0b", "c5b", "zkab", "zk-l", "zk-e", "zk-sw", "terminal", "tleb"], "Аксессуар СКУД"),
        (["tc-c3", "tc-c3", "tc-h3"], "Камера"),
        (["tc-p", "811", "a28"], "Аксессуар"),
        (["ds-n", "ds-h1"], "Регистратор"),
    ]
    for keywords, cat in checks:
        if any(k in nl for k in keywords):
            return cat
    return ""


row = 2
num = 1
for b in brands_tabs:
    items = missing.get(b, [])
    if not items:
        continue
    bd = brand_display[b]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=f"{bd} — {len(items)} позиций без характеристик")
    c.font = brand_font
    c.fill = brand_fill
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = brand_fill
        ws.cell(row=row, column=col).border = thin
    row += 1

    for name in items:
        nl = name.lower()
        cat = guess_category(nl)
        ws.cell(row=row, column=1, value=num).border = thin
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=bd).border = thin
        ws.cell(row=row, column=3, value=name).border = thin
        ws.cell(row=row, column=4, value=cat).border = thin
        row += 1
        num += 1

ws2 = wb.create_sheet("Сводка")
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 12

sum_headers = ["Бренд", "Всего", "С характер.", "Без характер.", "Покрытие"]
for col, h in enumerate(sum_headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin

brand_totals = {"dahua":1175,"hikvision":69,"hiwatch":52,"tiandy":85,"zkteco":136,"temid":66,"ezviz":14,"optimus":25,"belpark":1}
r = 2
total_all = 0
total_have = 0
for b in brands_tabs:
    bd = brand_display[b]
    total = brand_totals[b]
    miss = len(missing.get(b, []))
    have = total - miss
    total_all += total
    total_have += have
    pct = f"{round(100 * have / total)}%" if total > 0 else "—"
    ws2.cell(row=r, column=1, value=bd).border = thin
    ws2.cell(row=r, column=2, value=total).border = thin
    ws2.cell(row=r, column=3, value=have).border = thin
    ws2.cell(row=r, column=4, value=miss).border = thin
    ws2.cell(row=r, column=5, value=pct).border = thin
    for c in range(2, 6):
        ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    r += 1

ws2.cell(row=r, column=1, value="ИТОГО").font = Font(bold=True)
ws2.cell(row=r, column=1).border = thin
ws2.cell(row=r, column=2, value=total_all).font = Font(bold=True)
ws2.cell(row=r, column=2).border = thin
ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
ws2.cell(row=r, column=3, value=total_have).font = Font(bold=True)
ws2.cell(row=r, column=3).border = thin
ws2.cell(row=r, column=3).alignment = Alignment(horizontal="center")
total_miss = total_all - total_have
ws2.cell(row=r, column=4, value=total_miss).font = Font(bold=True)
ws2.cell(row=r, column=4).border = thin
ws2.cell(row=r, column=4).alignment = Alignment(horizontal="center")
ws2.cell(row=r, column=5, value=f"{round(100 * total_have / total_all)}%").font = Font(bold=True)
ws2.cell(row=r, column=5).border = thin
ws2.cell(row=r, column=5).alignment = Alignment(horizontal="center")

out = "Без_характеристик.xlsx"
wb.save(out)
print(f"Файл сохранён: {out}")
print(f"Позиций без характеристик: {num - 1}")
