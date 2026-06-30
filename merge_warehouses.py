import xlrd
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

sys.stdout.reconfigure(encoding="utf-8")


def open_sheet(filepath):
    with open(filepath, 'rb') as f:
        sig = f.read(4)
    is_zip = sig[:2] == b'PK'
    if is_zip:
        import tempfile, shutil
        if not filepath.endswith('.xlsx'):
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.close()
            shutil.copy2(filepath, tmp.name)
            filepath = tmp.name
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        nrows = ws.max_row or 0
        def cell_value(r, c):
            val = ws.cell(row=r+1, column=c+1).value
            return val if val is not None else ""
        return nrows, cell_value
    else:
        wb = xlrd.open_workbook(filepath)
        sh = wb.sheet_by_index(0)
        return sh.nrows, sh.cell_value


def parse_warehouse_items(filepath, warehouse_name):
    wb = xlrd.open_workbook(filepath)
    sh = wb.sheet_by_index(0)

    items = {}
    found = False
    for r in range(2, sh.nrows):
        v0 = sh.cell_value(r, 0)
        v1 = sh.cell_value(r, 1)

        if isinstance(v0, str) and v0.strip() == warehouse_name:
            found = True
            continue

        if found:
            if isinstance(v0, str) and v0.strip():
                name = v0.strip()
                qty = v1 if isinstance(v1, (int, float)) else 0

                # Check if this is the next warehouse header
                # Warehouse headers: check if the next few rows belong to this or it's a new section
                # A heuristic: if qty equals sum of next items, it's a warehouse header
                # Simpler: warehouse headers tend to have round or large numbers and short names
                # But actually the structure is: warehouse header, then items, then next warehouse header
                # We detect next warehouse by checking if a row's value matches known warehouse pattern
                # For now, let's collect until we hit a row that looks like another warehouse header

                # Items typically have specific product names with specs
                # Warehouse headers are short organizational names
                # Let's use: if the name doesn't contain digits or parentheses in first part,
                # and qty > sum_threshold, it might be a header
                # Actually simplest: just check if next row after a "header candidate" has items
                # But we need a different approach.

                # Let me just collect all items - we'll filter later
                items[name] = items.get(name, 0) + qty

            else:
                break  # empty row = end of warehouse section

    if not found:
        print(f"WARNING: Warehouse '{warehouse_name}' not found in {filepath}")

    return items


def find_next_warehouse(filepath, warehouse_name):
    """Find the row range for a warehouse and detect where it ends"""
    wb = xlrd.open_workbook(filepath)
    sh = wb.sheet_by_index(0)

    start_row = None
    for r in range(2, sh.nrows):
        v0 = str(sh.cell_value(r, 0)).strip()
        if v0 == warehouse_name:
            start_row = r
            break

    if start_row is None:
        return {}

    # Collect all items after warehouse header
    # End when we hit another warehouse header or end of file
    # We need to identify warehouse headers vs items
    # Strategy: read all rows, warehouse headers have their qty = sum of items below
    # But simpler: find ALL warehouse headers first, then extract items between them

    # First pass: find all warehouse-like headers
    # A warehouse header is at the same level as the found one
    all_warehouses = set()

    # Known warehouse patterns from our exploration
    # Let's find them by looking at the data structure
    # Items after a warehouse have product-like names (with specs, brands)
    # Warehouses have organizational names

    items = {}
    collecting = False
    for r in range(2, sh.nrows):
        v0 = str(sh.cell_value(r, 0)).strip()
        v1 = sh.cell_value(r, 1)

        if v0 == warehouse_name:
            collecting = True
            continue

        if collecting:
            if v0 == "" or v0 == "Итого":
                break

            # Try to detect if this is a new warehouse header
            # Check: is this name present as a warehouse in any of our known lists?
            # Simple heuristic: items usually have parentheses, digits, or long names with specs
            # Warehouses: short names without technical specs

            qty = v1 if isinstance(v1, (int, float)) else 0
            items[v0] = items.get(v0, 0) + qty

    return items


# Better approach: find ALL warehouse headers first
def get_all_warehouses(filepath):
    """Identify all warehouse headers by finding the hierarchical structure"""
    nrows, cell_value = open_sheet(filepath)

    warehouses = {}
    items_by_warehouse = {}

    rows = []
    for r in range(2, nrows):
        v0 = str(cell_value(r, 0)).strip()
        v1 = cell_value(r, 1) if isinstance(cell_value(r, 1), (int, float)) else 0
        rows.append((r, v0, v1))

    # Now find warehouse headers: go backwards, accumulate sums
    # A row is a warehouse header if its value equals the sum of all rows below it
    # until the next row whose value equals its own sum

    # Simpler brute force: try each row as potential warehouse header
    # Check if sum of rows below (until next potential header) equals this row's value
    i = 0
    warehouse_rows = []
    while i < len(rows):
        r, name, qty = rows[i]
        if qty > 0 and name:
            # Check sum of following rows
            running_sum = 0
            j = i + 1
            while j < len(rows):
                _, _, next_qty = rows[j]
                running_sum += next_qty
                if abs(running_sum - qty) < 0.01:
                    # This is a warehouse header!
                    warehouse_rows.append((i, r, name, qty, j))
                    break
                if running_sum > qty:
                    break
                j += 1
        i += 1

    return warehouse_rows, rows


def extract_items_between_warehouses(filepath, target_warehouse):
    """Extract items belonging to a specific warehouse"""
    warehouse_rows, all_rows = get_all_warehouses(filepath)

    # Find the target warehouse
    target_idx = None
    target_end = None
    for idx, (i, r, name, qty, end_j) in enumerate(warehouse_rows):
        if name == target_warehouse:
            target_idx = i
            target_end = end_j
            break

    if target_idx is None:
        print(f"WARNING: '{target_warehouse}' not found")
        return {}

    # But items between target and its end might contain sub-warehouses
    # For now, just collect all items between target_idx+1 and target_end
    # excluding any sub-warehouse headers

    # Get set of warehouse header indices
    wh_indices = set(i for i, r, name, qty, end_j in warehouse_rows)

    items = {}
    for k in range(target_idx + 1, target_end + 1):
        if k in wh_indices:
            continue  # skip sub-warehouse headers
        _, name, qty = all_rows[k]
        if name:
            items[name] = items.get(name, 0) + qty

    return items


def extract_all_items(filepath):
    """Extract all items from all warehouses, excluding warehouse headers and totals"""
    warehouse_rows, all_rows = get_all_warehouses(filepath)
    wh_indices = set(i for i, r, name, qty, end_j in warehouse_rows)
    items = {}
    for k, (r, name, qty) in enumerate(all_rows):
        if k in wh_indices:
            continue
        if name and name != "Итого" and qty > 0:
            items[name] = items.get(name, 0) + qty
    return items


# Main logic
def main():
    _dir = os.path.dirname(os.path.abspath(__file__))
    print("Parsing Минск (Основной АйТи Дистрибуция)...")
    minsk_items = extract_items_between_warehouses(
        os.path.join(_dir, "Минск.xls"),
        "Основной АйТи Дистрибуция"
    )
    print(f"  Найдено товаров: {len(minsk_items)}")

    print("Parsing Москва (Основной склад ЭлМир)...")
    moscow_items = extract_items_between_warehouses(
        os.path.join(_dir, "Москва.xls"),
        "Основной склад ЭлМир"
    )
    print(f"  Найдено товаров: {len(moscow_items)}")

    # Merge
    all_names = sorted(set(list(minsk_items.keys()) + list(moscow_items.keys())))
    print(f"  Всего уникальных наименований: {len(all_names)}")

    # Create output Excel
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Склады Минск и Москва"

    # Header styling
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Наименование", "Минск (АйТи Дистрибуция)", "Москва (ЭлМир)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, name in enumerate(all_names, 2):
        minsk_qty = minsk_items.get(name, 0)
        moscow_qty = moscow_items.get(name, 0)

        ws.cell(row=row_idx, column=1, value=name).border = thin_border
        ws.cell(row=row_idx, column=2, value=int(minsk_qty) if minsk_qty == int(minsk_qty) else minsk_qty).border = thin_border
        ws.cell(row=row_idx, column=3, value=int(moscow_qty) if moscow_qty == int(moscow_qty) else moscow_qty).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25

    # Auto-filter
    ws.auto_filter.ref = f"A1:C{len(all_names) + 1}"

    output_path = os.path.join(_dir, "Склады_Минск_Москва.xlsx")
    out_wb.save(output_path)
    print(f"\nФайл сохранён: {output_path}")
    print(f"Строк данных: {len(all_names)}")


if __name__ == "__main__":
    main()
