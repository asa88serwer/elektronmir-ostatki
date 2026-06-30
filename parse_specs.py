import csv
import re
import os
import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

DIR = os.path.dirname(os.path.abspath(__file__))

SPEC_LABELS = [
    ("category", "Категория"),
    ("device_type", "Тип устройства"),
    ("form", "Форм-фактор"),
    ("resolution", "Разрешение"),
    ("lens", "Объектив"),
    ("fov", "Угол обзора"),
    ("sensor", "Матрица"),
    ("ir", "Подсветка"),
    ("sensitivity", "Чувствительность"),
    ("wdr", "WDR"),
    ("codec", "Кодек"),
    ("fps", "Макс. к/с"),
    ("zoom", "Зум"),
    ("ip_rating", "Защита"),
    ("ik_rating", "Вандалозащита"),
    ("power", "Питание"),
    ("power_consumption", "Потребление"),
    ("housing", "Корпус"),
    ("mic", "Микрофон"),
    ("audio_io", "Аудио вх/вых"),
    ("alarm_io", "Тревога вх/вых"),
    ("sd_slot", "Карта памяти"),
    ("channels", "Каналов"),
    ("bandwidth", "Вх. поток"),
    ("rec_resolution", "Разр. записи"),
    ("storage", "Накопители"),
    ("video_out", "Видеовыходы"),
    ("poe_ports", "PoE-порты"),
    ("poe_power", "Мощность PoE"),
    ("analytics", "Аналитика"),
    ("ports", "Порты"),
    ("switch_capacity", "Комм. ёмкость"),
    ("switching_rate", "Скорость комм."),
    ("managed", "Управление"),
    ("surge", "Грозозащита"),
    ("verification", "Идентификация"),
    ("capacity", "Ёмкость"),
    ("interface", "Интерфейс"),
    ("display", "Дисплей"),
    ("temp_range", "Раб. температура"),
    ("dimensions", "Габариты"),
    ("weight", "Вес"),
    ("price", "РРЦ"),
]


def _read_xlsx_sheets(path, model_col, desc_col, price_col, header_row=1, skip_sheets=None):
    descs = {}
    prices = {}
    if not os.path.exists(path):
        return descs, prices
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        if skip_sheets and sn in skip_sheets:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            cells = [c.value for c in row]
            if len(cells) <= max(model_col, desc_col, price_col):
                continue
            model = str(cells[model_col] or "").strip()
            desc = str(cells[desc_col] or "").strip()
            if not model or not desc or len(model) < 3 or len(desc) < 10:
                continue
            if model not in descs:
                descs[model] = desc
            price_val = cells[price_col]
            if price_val and model not in prices:
                p = str(price_val).strip()
                if p and p != "None":
                    prices[model] = p
    wb.close()
    return descs, prices


def load_dahua_descriptions():
    descs = {}
    prices = {}
    csv_path = os.path.join(DIR, "прайсы", "Dahua_канальный_прайс.csv")
    if os.path.exists(csv_path):
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f, delimiter=";"):
                if len(row) >= 2 and row[0].strip() and row[1].strip():
                    art = row[0].strip()
                    descs[art] = row[1].strip()
                    if len(row) >= 3 and row[2].strip():
                        prices[art] = row[2].strip()

    xlsx_path = os.path.join(DIR, "прайсы", "Dahua_проектный_прайс_15.06.2026.xlsx")
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            art = str(ws.cell(r, 1).value or "").strip()
            desc = str(ws.cell(r, 2).value or "").strip()
            price_val = ws.cell(r, 3).value
            if art and desc and art not in descs:
                descs[art] = desc
                if price_val:
                    prices[art] = str(price_val)

    po_path = os.path.join(DIR, "прайсы", "Dahua_CyberCity_ПО.csv")
    if os.path.exists(po_path):
        with open(po_path, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f, delimiter=";"):
                if len(row) >= 3 and row[1].strip() and row[2].strip():
                    art = row[1].strip()
                    if art not in descs:
                        descs[art] = row[2].strip()
                    if len(row) >= 4 and row[3].strip():
                        prices[art] = row[3].strip()

    return descs, prices


def load_hikvision_descriptions():
    path = os.path.join(DIR, "прайсы", "Hikvision_CCTV_прайс.xlsx")
    return _read_xlsx_sheets(path, model_col=1, desc_col=3, price_col=5,
                             header_row=3, skip_sheets={"СОДЕРЖАНИЕ"})


def load_hiwatch_descriptions():
    path = os.path.join(DIR, "прайсы", "HiWatch_CCTV_прайс.xlsx")
    return _read_xlsx_sheets(path, model_col=2, desc_col=3, price_col=6, header_row=2)


def load_tiandy_descriptions():
    descs = {}
    prices = {}
    rrc_path = os.path.join(DIR, "прайсы", "Tiandy_прайс_РРЦ.xlsx")
    if os.path.exists(rrc_path):
        d, p = _read_xlsx_sheets(rrc_path, model_col=3, desc_col=6, price_col=4, header_row=1)
        descs.update(d)
        prices.update(p)
    proj_path = os.path.join(DIR, "прайсы", "Tiandy_проектный_прайс.xlsx")
    if os.path.exists(proj_path):
        d, p = _read_xlsx_sheets(proj_path, model_col=1, desc_col=4, price_col=2, header_row=1)
        for k, v in d.items():
            if k not in descs:
                descs[k] = v
        for k, v in p.items():
            if k not in prices:
                prices[k] = v
    return descs, prices


def load_temid_prices():
    prices = {}
    path = os.path.join(DIR, "прайсы", "Temid_РРЦ_прайс.xlsx")
    if not os.path.exists(path):
        return prices
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or len(row) < 3:
            continue
        model = str(row[1] or "").strip()
        price_val = row[2]
        if model and price_val:
            p = str(price_val).strip()
            if p and p not in ("None", "по запросу"):
                prices[model] = p
    wb.close()
    return prices


def parse_specs(desc):
    specs = {}
    dl = desc.lower()

    if "коммутатор" in dl or "свитч" in dl or "switch" in dl:
        specs["category"] = "Коммутатор"
    elif "роутер" in dl or "точка доступа" in dl:
        specs["category"] = "Сетевое"
    elif "крепление" in dl or "кронштейн" in dl or ("монтаж" in dl and "видеокамер" not in dl) or "адаптер" in dl:
        specs["category"] = "Аксессуар"
    elif "видеокамер" in dl or "ip-камер" in dl or "камера" in dl:
        specs["category"] = "Камера"
    elif "видеорегистратор" in dl or "nvr" in dl or "dvr" in dl:
        specs["category"] = "Регистратор"
    elif "монитор" in dl or "дисплей" in dl or "панель" in dl:
        specs["category"] = "Монитор"
    elif "лицензи" in dl or "программн" in dl:
        specs["category"] = "ПО"
    elif "жёстк" in dl or "жестк" in dl or "ssd" in dl or "накопител" in dl:
        specs["category"] = "Хранение"
    else:
        specs["category"] = "Другое"

    forms = []
    if "купольн" in dl or "турель" in dl:
        forms.append("Купольная")
    if "цилиндрическ" in dl:
        forms.append("Цилиндрическая")
    if "ptz" in dl or "поворотн" in dl or "скоростн" in dl:
        forms.append("PTZ")
    if "fisheye" in dl or "панорамн" in dl:
        forms.append("Панорамная")
    if "мини" in dl and "поворотн" in dl:
        forms = ["Мини-поворотная"]
    if forms:
        specs["form"] = ", ".join(forms)

    m = re.search(r"(\d+)\s*[Мм][Пп]", desc)
    if m:
        specs["resolution"] = m.group(1) + "Мп"
    else:
        m = re.search(r"(\d{3,4})\s*[×xXх]\s*(\d{3,4})", desc)
        if m:
            w, h = int(m.group(1)), int(m.group(2))
            mp = round(w * h / 1_000_000, 1)
            if mp >= 1:
                specs["resolution"] = f"{mp}Мп"

    m = re.search(r"объектив\s+([^\s,;]+)", desc, re.I)
    if not m:
        m = re.search(r"(\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?)\s*мм", desc)
    if m:
        val = m.group(1) if m.lastindex == 1 else m.group(0)
        val = val.replace("объектив", "").strip()
        if not val.endswith("мм") and re.match(r"[\d.-]+$", val):
            val += "мм"
        specs["lens"] = val

    m = re.search(r"(?:ИК|IR|LED|EXIR)[-\s]*(?:подсветк[аи])?\s*(?:до\s+)?(\d+)\s*м(?:\b|[.,;])", desc)
    if m:
        specs["ir"] = "до " + m.group(1) + "м"

    m = re.search(r"(?:чувствительность|чувств)[:\s]*([\d.,]+)\s*л[кx]", desc, re.I)
    if not m:
        m = re.search(r"([\d.,]+)\s*л[кx]", desc)
    if m:
        specs["sensitivity"] = m.group(1) + " лк"

    m = re.search(r"WDR.*?(\d+)\s*дБ", desc)
    if not m:
        m = re.search(r"(\d+)\s*дБ.*?WDR", desc)
    if not m and "dwdr" in dl:
        specs["wdr"] = "DWDR"
    elif m:
        specs["wdr"] = m.group(1) + " дБ"

    codecs = []
    for c in ["S+265", "H.265+", "H.265", "H.264+", "H.264", "MJPEG"]:
        if c in desc:
            codecs.append(c)
    if codecs:
        specs["codec"] = ", ".join(codecs[:3])

    m = re.search(r"(IP\d+)", desc)
    if m:
        specs["ip_rating"] = m.group(1)

    pwr = []
    if "PoE+" in desc or "802.3at" in desc:
        pwr.append("PoE+")
    elif "PoE" in desc or "802.3af" in desc:
        pwr.append("PoE")
    if "12В" in desc or "12V" in desc or "12 DC" in desc or "DC12" in desc:
        pwr.append("12В")
    if "24В" in desc or "24V" in desc:
        pwr.append("24В")
    if "48В" in desc or "48V" in desc:
        pwr.append("48В")
    if pwr:
        specs["power"] = ", ".join(pwr)

    m = re.search(r"корпус:\s*(\S+)", desc, re.I)
    if m:
        val = m.group(1).rstrip(",.")
        specs["housing"] = val

    m = re.search(r"(\d+)\s*-?\s*канальн", desc)
    if not m:
        m = re.search(r"(\d+)\s*канал", desc)
    if m:
        specs["channels"] = m.group(1)

    m = re.search(r'(1/[\d.]+")[?\s]*CMOS', desc)
    if not m:
        m = re.search(r"(1/[\d.]+''\s*CMOS)", desc)
    if not m:
        m = re.search(r"(1/[\d.]+\")\s*(?:Progressive\s+Scan\s+)?CMOS", desc)
    if m:
        specs["sensor"] = m.group(1).rstrip("'\"") + "\" CMOS"

    if "встроенный микрофон" in dl or "встр. микрофон" in dl:
        specs["mic"] = "Встроенный"
    elif "микрофон" in dl:
        specs["mic"] = "Да"

    m = re.search(r"[Вв]ход(?:ящий|\.?) поток[:\s]+(?:до\s+)?([\d]+)\s*Мбит", desc)
    if not m:
        m = re.search(r"битрейт[:\s]+([\d]+)\s*Мбит", desc)
    if m:
        specs["bandwidth"] = m.group(1) + " Мбит/с"

    m = re.search(r"разреш(?:ение)?\s*записи[:\s]+до\s+(\d+\s*Мп)", desc, re.I)
    if m:
        specs["rec_resolution"] = m.group(1)

    m = re.search(r"(?:накопители|HDD)[:\s]*(\d+\s*SATA[^,;]*)", desc, re.I)
    if not m:
        m = re.search(r"(\d+)\s*(?:HDD|SATA)", desc)
        if m:
            specs["storage"] = m.group(1) + " HDD"
    if m and "storage" not in specs:
        specs["storage"] = m.group(1).strip()

    vouts = []
    for vm in re.finditer(r"(\d+)\s*(HDMI|VGA)", desc):
        vouts.append(vm.group(1) + " " + vm.group(2))
    if vouts:
        specs["video_out"] = ", ".join(vouts)

    m = re.search(r"(\d+)\s*(?:RJ45|портов?)[^,]*PoE", desc)
    if m:
        specs["poe_ports"] = m.group(1)

    m = re.search(r"суммарно\s+до\s+(\d+)\s*Вт", desc, re.I)
    if m:
        specs["poe_power"] = m.group(1) + " Вт"
    else:
        m2 = re.search(r"PoE[^,]*до\s+(\d+)\s*Вт", desc)
        if m2:
            specs["poe_power"] = m2.group(1) + " Вт"

    analytics = []
    if "smd plus" in dl or "smd+" in dl:
        analytics.append("SMD Plus")
    elif "smd" in dl:
        analytics.append("SMD")
    if "acusense" in dl:
        analytics.append("AcuSense")
    if "colorvu" in dl:
        analytics.append("ColorVu")
    if "распознавание лиц" in dl or "классификация" in dl:
        analytics.append("Распознавание лиц")
    if "охрана периметра" in dl or "пересечени" in dl:
        analytics.append("Охрана периметра")
    if "обнаружение людей" in dl or "человек" in dl:
        analytics.append("Детекция людей")
    if "обнаружение транспорт" in dl or "тс" in dl.split():
        analytics.append("Детекция ТС")
    if analytics:
        specs["analytics"] = ", ".join(dict.fromkeys(analytics))

    m = re.search(r"[Пп]орты:\s*(.+?)(?:,\s*[Кк]оммутац|,\s*мощность|,\s*питание|$)", desc)
    if m and ("коммутатор" in dl or "свитч" in dl):
        specs["ports"] = m.group(1).strip().rstrip(",")

    m = re.search(r"[Кк]оммутационная\s+[её]мкость:\s*([\d.,]+\s*Гбит/с)", desc)
    if m:
        specs["switch_capacity"] = m.group(1)

    m = re.search(r"[Сс]корость коммутации[^:]*:\s*([\d.,]+\s*Мпак/с)", desc)
    if m:
        specs["switching_rate"] = m.group(1)

    if "неуправляем" in dl:
        specs["managed"] = "Неуправляемый"
    elif "облачн" in dl:
        specs["managed"] = "Облачное"
    elif "управляем" in dl:
        specs["managed"] = "Управляемый"

    m = re.search(r"грозозащита:\s*до\s+(\d+\s*кВ)", desc, re.I)
    if m:
        specs["surge"] = m.group(1)

    m = re.search(r"(?:рабочая\s+)?температура[:\s]*([−\-]?\d+\s*[°.]*\s*[CС]?\s*\.{0,3}\s*[+]?\d+\s*[°.]*\s*[CС]?)", desc, re.I)
    if not m:
        m = re.search(r"([−\-]\d+\s*[°.…]+\s*[+]?\d+\s*°?\s*[CС])", desc)
    if m:
        specs["temp_range"] = m.group(1).strip()

    m = re.search(r"угол\s+обзора[:\s]*([\d.,]+\s*°?)", desc, re.I)
    if not m:
        m = re.search(r"([\d.]+)\s*°", desc)
    if m:
        specs["fov"] = m.group(1).strip() + ("°" if "°" not in m.group(0) else "")

    m = re.search(r"(\d+)\s*к/с", desc)
    if not m:
        m = re.search(r"(\d+)\s*fps", desc, re.I)
    if m:
        specs["fps"] = m.group(1) + " к/с"

    m = re.search(r"зум\s*(\d+)[xXхХ]", desc, re.I)
    if not m:
        m = re.search(r"(\d+)[xXхХ]\s*(?:оптич|zoom)", desc, re.I)
    if m:
        specs["zoom"] = m.group(1) + "x"

    m = re.search(r"(IK\d+)", desc)
    if m:
        specs["ik_rating"] = m.group(1)

    m = re.search(r"(\d+(?:[.,]\d+)?)\s*Вт\s*(?:макс|max)?", desc, re.I)
    if m and "poe_power" not in specs:
        specs["power_consumption"] = m.group(1) + " Вт"

    m = re.search(r"аудио(?:вход|вх)[/\\](?:выход|вых)[:\s]*(\d+/\d+)", desc, re.I)
    if not m:
        m = re.search(r"аудио[:\s]*(\d+\s*вх[./]\s*\d+\s*вых)", desc, re.I)
    if m:
        specs["audio_io"] = m.group(1)

    m = re.search(r"тревожн[а-я]*\s+вх[./а-я]*/вых[а-я]*[:\s]*(\d+/\d+)", desc, re.I)
    if m:
        specs["alarm_io"] = m.group(1)

    m = re.search(r"(?:microSD|micro\s*SD|SD)[^;,]*?до\s+(\d+)\s*(?:Гб|GB)", desc, re.I)
    if not m:
        m = re.search(r"слот\s+для\s+(?:micro)?SD", desc, re.I)
    if m:
        specs["sd_slot"] = ("до " + m.group(1) + " ГБ") if m.lastindex and m.lastindex >= 1 else "Да"

    m = re.search(r"(\d+(?:[.,]\d+)?)\s*кг", desc)
    if m:
        specs["weight"] = m.group(1) + " кг"

    m = re.search(r"(\d+(?:[.,]\d+)?\s*[×xXх]\s*\d+(?:[.,]\d+)?(?:\s*[×xXх]\s*\d+(?:[.,]\d+)?)?)\s*(?:мм|mm)", desc)
    if m:
        specs["dimensions"] = m.group(1) + " мм"

    return specs


def _normalize_model(s):
    s = re.sub(r"\s+", "", s)
    s = s.split("(")[0].strip()
    return s.upper()


def _base_model(s):
    m = re.match(r"((?:TC-|DS-|DH-)?[A-Za-z0-9-]+)", s.strip())
    return m.group(1).upper() if m else s.upper()


def _match_article_to_product(art, product_names, brand_prefix):
    for pname in product_names:
        if art in pname:
            return pname
    art_norm = _normalize_model(art)
    if len(art_norm) < 4:
        return None
    for pname in product_names:
        pl = pname.lower()
        if pl.startswith(brand_prefix):
            rest = pname[len(brand_prefix):].strip()
            pname_norm = _normalize_model(rest)
            if pname_norm == art_norm:
                return pname
            if art_norm in pname_norm or pname_norm in art_norm:
                return pname
    art_base = _base_model(art)
    if len(art_base) >= 8:
        candidates = []
        for pname in product_names:
            pl = pname.lower()
            if pl.startswith(brand_prefix):
                rest = pname[len(brand_prefix):].strip()
                if _base_model(rest) == art_base:
                    candidates.append(pname)
        if len(candidates) == 1:
            return candidates[0]
    return None


def _load_web_specs(existing_specs):
    added = 0
    for fname in ["web_specs_tinko.json", "web_specs_ozon.json"]:
        fpath = os.path.join(DIR, fname)
        if not os.path.exists(fpath):
            continue
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        for name, sp in data.items():
            if name not in existing_specs and sp:
                existing_specs[name] = sp
                added += 1
    return added


def _specs_from_product_names(product_names, existing_specs):
    added = 0
    brand_tabs = ["dahua", "hikvision", "hiwatch", "tiandy", "zkteco", "temid",
                  "ezviz", "optimus", "belpark"]
    for name in product_names:
        if name in existing_specs:
            continue
        nl = name.lower()
        if not any(nl.startswith(b) for b in brand_tabs):
            continue
        specs = {}
        if re.search(r"(\d+)\s*[Мм][Пп]", name):
            specs["resolution"] = re.search(r"(\d+)\s*[Мм][Пп]", name).group(1) + "Мп"
        elif re.search(r"(\d)[KkКк]", name):
            k = int(re.search(r"(\d)[KkКк]", name).group(1))
            mp_map = {2: "2Мп", 3: "3Мп", 4: "4Мп", 5: "5Мп", 8: "8Мп"}
            if k in mp_map:
                specs["resolution"] = mp_map[k]
        m = re.search(r"(\d+(?:\.\d+)?)\s*(?:мм|mm)", name, re.I)
        if m:
            specs["lens"] = m.group(1) + "мм"
        if "ipc" in nl or "ip-камер" in nl or "cs-" in nl:
            specs["category"] = "Камера"
        elif "nvr" in nl or "dvr" in nl or "регистр" in nl:
            specs["category"] = "Регистратор"
        elif "poe" in nl or "коммутат" in nl:
            specs["category"] = "Коммутатор"
        if "poe" in nl:
            m = re.search(r"(\d+)\s*Вт", name)
            if m:
                specs["poe_power"] = m.group(1) + " Вт"
            m = re.search(r"(\d+)\s*(?:x|х)?\s*(?:FE|GE|порт)", name, re.I)
            if m:
                specs["ports"] = m.group(0).strip()
        if specs and len(specs) >= 1:
            existing_specs[name] = specs
            added += 1
    return added


def build_specs_map(product_names):
    specs_map = {}
    total_matched = 0

    loaders = [
        ("Dahua", load_dahua_descriptions, "dahua "),
        ("Hikvision", load_hikvision_descriptions, "hikvision "),
        ("HiWatch", load_hiwatch_descriptions, "hiwatch "),
        ("Tiandy", load_tiandy_descriptions, "tiandy "),
    ]

    for brand, loader, prefix in loaders:
        descs, prices = loader()
        if not descs:
            continue
        matched = 0
        for art, desc in descs.items():
            name = _match_article_to_product(art, product_names, prefix)
            if name and name not in specs_map:
                sp = parse_specs(desc)
                if art in prices:
                    sp["price"] = prices[art]
                if sp:
                    specs_map[name] = sp
                    matched += 1
        total_matched += matched
        print(f"  {brand}: {matched} из {len(descs)} артикулов сопоставлено")

    temid_prices = load_temid_prices()
    if temid_prices:
        temid_matched = 0
        for model, price in temid_prices.items():
            name = _match_article_to_product(model, product_names, "zkteco ")
            if name and name not in specs_map:
                specs_map[name] = {"category": "СКУД", "price": price}
                temid_matched += 1
        print(f"  ZKTeco/Temid: {temid_matched} цен сопоставлено (без описаний)")
        total_matched += temid_matched

    web_added = _load_web_specs(specs_map)
    if web_added:
        print(f"  Из веб-источников: {web_added} доп. товаров")

    fallback = _specs_from_product_names(product_names, specs_map)
    if fallback:
        print(f"  Из названий: {fallback} доп. товаров")

    print(f"  Итого: {len(specs_map)} товаров с характеристиками")
    return specs_map


if __name__ == "__main__":
    from merge_warehouses import extract_all_items
    minsk = extract_all_items(os.path.join(DIR, "Минск.xls"))
    moscow = extract_all_items(os.path.join(DIR, "Москва.xls"))
    all_names = sorted(set(list(minsk.keys()) + list(moscow.keys())))
    print(f"Всего товаров на складах: {len(all_names)}")

    specs = build_specs_map(all_names)
    print(f"\nИтого с характеристиками: {len(specs)}")

    brands = {}
    for name in specs:
        b = name.split()[0] if name.split() else "?"
        brands[b] = brands.get(b, 0) + 1
    for b, c in sorted(brands.items(), key=lambda x: -x[1]):
        print(f"  {b}: {c}")

    for name in list(specs.keys())[:5]:
        print(f"\n{name}:")
        print(f"  {json.dumps(specs[name], ensure_ascii=False)}")
